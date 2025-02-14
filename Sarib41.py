from tkinter import *
import pandas as pd
import os.path
from openpyxl import Workbook, load_workbook
from datetime import datetime,date
from tkinter import *
Time = datetime.now().strftime('%H:%M:%S')
time = datetime.now().strftime('%H%M%S')
Date =date.today()
if os.path.isfile('Facebook.xlsx') == True:
    wb = load_workbook('Facebook.xlsx')
else:
    wb = Workbook()
    ws = wb.active
    ws.title = 'All Users'
    ws.append(['Indexes','Name','Password','Resources','About','Address','Nick Name','Education','Work','Logged-In','Notification','Messages','Friend_Requests','Comments'])
    wb.create_sheet('Friends')
    wb.create_sheet('Messages')
    wb.create_sheet('Friend_Requests')
    wb.create_sheet('Posts')
    wb.create_sheet('Ban')
    wb.create_sheet('Block')
    wb.create_sheet('Page')
    wb.create_sheet('Page Post')
    wb.create_sheet('Ban_Post')
wb.save('Facebook.xlsx')
class Facebook:
    error='                         '
    search_error = ''
    message_error = ''
    cmnt='abc'
    Cmnt_Difference = 0
    i=1
    posted = ''
    searched = ''
    rqst_indx = 3
    counter = 1
    wb['Friends']['A1'] = 'Indexes'
    wb['Messages']['A1'] = 'Indexes'
    wb['Friend_Requests']['A1'] = 'Indexes'
    wb['Friends']['B1'] = 'Name'
    wb['Friend_Requests']['B1'] = 'Name'
    wb['Messages']['B1'] = 'Person1'
    wb['Messages']['C1'] = 'Person2'
    wb.save('Facebook.xlsx')
    row = wb['All Users']['J2'].value
    if row == None:
        From = wb['All Users'].cell(row=1, column=2).value
    else:
        From = wb['All Users'].cell(row=row + 1, column=2).value
    def __init__(self,id,set):
        def sign_up_call():
            ID.Pre_Sign_up(id)
        Facebook.id=id
        Facebook.Sett = set
        if wb['All Users']['J2'].value == None:
            root = Tk()
            root.title('facebook')
            root.attributes('-fullscreen', True)
            Label(root,text='facebook',bg='blue',fg='white',font='Arial 25 bold').place(x=80,y=170)
            Label(root,text='By Sarib Gardazi',fg='red',font='Arial 12 bold').place(x=250,y=180)
            Label(root,text='Shaping those who will shape the future',fg='blue',font='Arial 20 bold').place(x=100,y=240)
            frame = Frame(root,bg='#fffcfc',borderwidth=4)
            frame.place(x=700,y=180,width=400,height=350)
            username = Label(root, text='Username',fg='red',bg='#fffcfc', font='Simple 12 bold')
            password = Label(root, text="Password",fg='red',bg='#fffcfc', font='Simple 12 bold')
            self.name = Entry(root, width=30,bg='#fffcfc', textvariable=StringVar(),borderwidth=3)
            self.Pasword = Entry(root, width=30,bg='#fffcfc', textvariable=StringVar(),show='*',borderwidth=3)
            username.place(x=750, y=200, height=30)
            password.place(x=750, y=250, height=30)
            self.name.place(x=850, y=200, height=30)
            self.Pasword.place(x=850, y=250, height=30)
            space = Label(root,text=Facebook.error,fg='red',bg='#fffcfc',font='Arial 12 italic')
            space.place(x=800,y=280,height=50)
            Button(root, text='Log In', width=30, bg='blue', fg='#fffcfc', font='Simple 12 bold',command=self.LOG_IN_CALL).place(x=750, y=330, height=50)
            Label(text='OR',fg='black',bg='white',font='Arial 15 bold').place(x=880,y=390)
            Button(root, text='Create Account', width=27, bg='green', fg='#fffcfc', font='Simple 14 bold',command=sign_up_call).place(x=750, y=430, height=50)
            root.mainloop()
        else:
            Facebook.MainPage(self)
    def LOG_IN_CALL(self):
        self.name1 = self.name.get()
        self.password1 = self.Pasword.get()
        self.LOG_IN()
    def LOG_IN_SIGN(self):
        self.name1 = ID.name2
        self.password1 = ID.Pasword2
        self.LOG_IN()
    def LOG_IN(self):
        name = self.name1
        password = self.password1
        if name in ID.Name:
            index = ID.Name.index(name)
            if password == wb['All Users']['C'+ str(index+2)].value:
                wb['All Users']['J2']= index + 1
                wb.save('Facebook.xlsx')
                Facebook.row = wb['All Users']['J2'].value
                Facebook.From = wb['All Users'].cell(row=Facebook.row + 1, column=2).value
                wb.save('Facebook.xlsx')
                Facebook.error = '                         '
                Facebook.search_error = ''
                Facebook.message_error = ''
                Facebook.cmnt = 'abc'
                Facebook.Cmnt_Difference = 0
                Facebook.i = 1
                Facebook.posted = ''
                Facebook.searched = ''
                Facebook.rqst_indx = 3
                self.MainPage()
            else:
                Facebook.error = 'Invalid Password'
                Facebook(Facebook.id)
        else:
            Facebook.error = 'Invalid User Name'
            Facebook(Facebook.id,Facebook.Sett)
        Facebook(ID(),Setting())
    def Pre_Search(self):
        frnd_list = []
        a=3
        while wb['Friends'].cell(row=Facebook.row+1,column=a).value != None:
            frnd_list.append(wb['Friends'].cell(row=Facebook.row+1,column=a).value)
            a+=1
        root = Tk()
        root.title('facebook')
        root.attributes('-fullscreen',True)
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30, y=30)
        hint = Label(root,text = 'Type your search key here ',fg='black',font='Arial 12 bold')
        hint.place(x=500,y=170)
        Label(root,text= Facebook.search_error,fg='red').place(x=510,y=150)
        Facebook.search = Entry(root,width=30, textvariable=StringVar())
        Facebook.search.place(x=510,y=210,height=30)
        search_button = Button(root,width=10,height=2,text='Search',bg='blue',fg='white',font='Simple 10 bold',command=self.Search1).place(x=550,y=250)
        root.mainloop()
    def Search1(self):
        Facebook.frnd = Facebook.search.get()
        self.Search()
    def Profile(self):
        Facebook.frnd = wb['All Users'].cell(row=Facebook.row+1, column=2).value
        self.Search()
    def Search(self):
        if Facebook.frnd in ID.Name:
            ind = ID.Name.index(Facebook.frnd) + 2
            a = 1
            check = True
            while wb['Block'].cell(row=ind, column=a).value != None:
                if wb['Block'].cell(row=ind, column=a).value == Facebook.From:
                    check = False
                a += 1
            if check:
                col = 2
                root = Tk()
                root.title('facebook')
                root.attributes('-fullscreen',True)
                Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30,y=30)
                Profile = Label(root,text='Profile',bg='blue',fg='white',font='Simple 14 bold')
                Profile.place(x=600,y=50)
                lst = ['b1','','b2','b3','b4','b5','b6','b7']
                lst1= [1,2,3,4,5,6,7,8]
                indicator = 3
                while wb['All Users'].cell(row=1, column=col + 5).value != None:
                    if wb['All Users'].cell(row=1, column=col).value != 'Password':
                        lst[col-2] = Label(root,text = wb['All Users'].cell(row=1, column=col).value,fg='brown',font='Allstar 12 bold')
                        lst1[col-2] = Label(root,text = wb['All Users'].cell(row=ind, column=col).value,fg='blue',font='Allstar 12 bold')
                        lst[col-2].place(x=500 , y = indicator*30)
                        lst1[col-2].place(x=700,y = indicator*30)
                        indicator+=1
                    col += 1
                i = 1
                j=1
                Post = Label(root, text='Posts',fg='blue', font='Simple 14 bold')
                Post.place(x=500,y=350)
                Label(root,text=Facebook.searched,fg='red',font='Arial 8 italic').place(x=600,y=350)
                Button(root,text='Friend Request',bg='blue',fg='white',font='Arial 12 bold',command=self.Make_Frnd).place(x=600,y=310)
                Button(root, text='  Friends  ', bg='white', fg='blue', font='Arial 12 bold',command=self.Show_frnds).place(x=500, y=310)
                while wb['Posts']['C' + str(i)].value != None:
                    check = True
                    if wb['Posts']['B' + str(i)].value == wb['All Users'].cell(row=ind, column=2).value:
                        a = 1
                        while wb['Ban_Post'].cell(row=i, column=a).value != None:
                            if wb['Ban_Post'].cell(row=i, column=a).value == Facebook.From:
                                check = False
                            a += 1
                        if check:
                            Label(root,text=wb['Posts']['B' + str(i)].value,fg='red',font='Arial 13 bold').place(x=500,y=(j*80)+300)
                            Label(root,text=wb['Posts']['C' + str(i)].value,fg='blue',font='Arial 12 bold').place(x=580,y=(j*80)+300)
                            Label(root,text=wb['Posts']['D' + str(i)].value,fg='black',font='Arial 12 bold').place(x=500,y=(j*80)+330)
                            j += 1
                    i+=1
                root.mainloop()
        else:
            print(1)
            check=False
            for i in range(1,20000):
                lst=[]
                if wb['Posts']['D'+str(i)].value != None:
                    txt = wb['Posts']['D'+str(i)].value
                    lst = txt.split()
                if self.frnd in lst:
                    print(2)
                    root = Tk()
                    root.attributes('-fullscreen',True)
                    Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30, y=30)
                    Label(root, text=wb['Posts']['B' + str(i)].value, fg='red', font='Arial 18 bold').place(x=200, y=220)
                    Label(root, text=wb['Posts']['C' + str(i)].value, fg='blue', font='Arial 17 bold').place(x=300, y=220)
                    Label(root, text=wb['Posts']['D' + str(i)].value, fg='black', font='Arial 17 bold').place(x=200,y=280)
                    root.mainloop()
                    check=True
                elif wb['Posts']['D'+str(i)].value == None:
                    break
            if check == False:
                Facebook.search_error = 'No user or post found...'
                self.Pre_Search()
        # else:
        #     Facebook.search_error = 'No user or post found...'
        #     self.Pre_Search()
    def Show_frnds(self):
        frnd_lst = []
        ind = ID.Name.index(Facebook.frnd) + 2
        for i in range(3,2000):
            if wb['Friends'].cell(row=ind,column=i).value == None:
                break
            else:
                frnd_lst.append(wb['Friends'].cell(row=ind,column=i).value)
        root = Tk()
        root.attributes('-fullscreen',True)
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30, y=30)
        Label(root,text='Friends',fg='blue',font='Arial 20 bold').place(x=200,y=100)
        for i in range(len(frnd_lst)):
            Label(root,text=frnd_lst[i],fg='brown',font='Arial 18 bold').place(x=250,y=(i*60)+150)
        root.mainloop()
    def Make_Frnd(self):
        row=ID.Name.index(self.frnd)+2
        frnd_req_list = []
        frnd_list = []
        a = 3
        while wb['Friend_Requests'].cell(row=row, column=a).value != None:
            frnd_req_list.append(wb['Friend_Requests'].cell(row=row, column=a).value)
            a += 1
        a=3
        while wb['Friends'].cell(row=row, column=a).value != None:
            frnd_list.append(wb['Friends'].cell(row=row, column=a).value)
            a += 1
        if Facebook.From in frnd_list:
            Facebook.searched = 'You are already Friends of each other.'
            self.Search()
        elif Facebook.From in frnd_req_list:
            Facebook.searched = 'Your Request Already sent. You cant sent again'
            self.Search()
        elif self.frnd in ID.Name and self.frnd != Facebook.From:
            frnd_index = ID.Name.index(self.frnd) + 2
            col = Facebook.Finding_index(self,wb,frnd_index,'C','Friend_Requests')
            wb['Friend_Requests'].cell(row=frnd_index,column=col,value=self.From)
            wb['All Users'].cell(row=frnd_index,column=11).value += 1
            wb['All Users'].cell(row=frnd_index, column=13).value += 1
            wb.save('Facebook.xlsx')
            Facebook.searched = 'Request sent'
            self.Search()
        else:
            Facebook.searched = 'Its your own ID....'
            self.Search()
    def check_friends(self):
        root = Tk()
        root.attributes('-fullscreen',True)
        Label(root,text='Friend Requests',bg='blue',fg='white',font='Arial 14 bold').place(x=500,y=50)
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30, y=30)
        if wb['Friend_Requests'].cell(row=self.row+1,column=self.rqst_indx).value != None:
            Label(root,text=wb['Friend_Requests'].cell(row=self.row+1,column=self.rqst_indx).value,fg='blue',font='Arial 14 bold').place(x=500,y=240)
            Label(root,text='sent you a Friend Request',font='Arial 12 bold').place(x=560,y=242)
            Button(root,text='Accept',bg='blue',fg='white',font='Arial 14 bold',command=self.Frnd_Accepted).place(x=700,y=300)
            Button(root,text='Decline',bg='white',fg='blue',font='Arial 14 bold',command=self.Declined).place(x=500,y=300)
        root.mainloop()
    def Frnd_Accepted(self):
                frnd = wb['Friend_Requests'].cell(row=self.row + 1, column=self.rqst_indx).value
                frnd_index = ID.Name.index(frnd) + 2
                col = Facebook.Finding_index(self,wb,frnd_index,'C','Friends')
                wb['Friends'].cell(row=frnd_index, column=col, value=self.From)
                col = Facebook.Finding_index(self, wb,self.row+1, 'C','Friends')
                wb['Friends'].cell(row=self.row+1, column=col, value=frnd)
                wb['Friend_Requests'].cell(row=self.row + 1, column=self.rqst_indx).value = None
                wb['All Users']['M'+str(self.row + 1)].value -=1
                wb.save('Facebook.xlsx')
                self.rqst_indx += 1
                self.check_friends()
    def Declined(self):
        wb['Friend_Requests'].cell(row=self.row + 1, column=self.rqst_indx).value = None
        wb['All Users']['M' + str(self.row + 1)].value -= 1
        wb.save('Facebook.xlsx')
        self.rqst_indx += 1
        self.check_friends()
    def Pre_Messages(self):
        lst=[]
        for i in range(1,2000):
            if wb['Messages'].cell(row=i,column=2).value == Facebook.From:
                lst.append(wb['Messages'].cell(row=i,column=3).value)
            elif wb['Messages'].cell(row=i,column=3).value == Facebook.From:
                lst.append(wb['Messages'].cell(row=i,column=2).value)
            elif wb['Messages'].cell(row=i,column=3).value == None:
                break
        root=Tk()
        root.attributes('-fullscreen',True)
        for i in range(len(lst)):
            Label(root,text=lst[i],fg='blue',font='Harrington 20 bold').place(x=150,y=(50*i)+200)
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30,y=30)
        Label(root, text=Facebook.message_error, fg='red').place(x=620, y=130)
        self.search = Entry(root, width=30, textvariable=StringVar())
        self.search.insert(0,'Type search key here')
        def click(event):
            self.search.delete(0,'end')
        self.search.bind('<Button-1>',click)
        self.search.place(x=610, y=150, height=30)
        Label(root,text='Recently Chats',fg='Purple',font='Arial 20 bold').place(x=150,y=150)
        Button(root,text='Search', bg='blue', fg='white', font='Simple 10 bold',command=self.Messaging).place(x=800, y=150)
        root.mainloop()
    def Messaging(self):
        frnd = self.search.get()
        if frnd not in ID.Name:
            Facebook.message_error = 'No user found!'
            self.Pre_Messages()
        elif self.From == frnd:
            Facebook.message_error = 'You Cant send message to your own id'
            self.Pre_Messages()
        else:
            ind = ID.Name.index(frnd) + 2
            a = 1
            check = True
            while wb['Block'].cell(row=ind, column=a).value != None:
                if wb['Block'].cell(row=ind, column=a).value == Facebook.From:
                    check = False
                a += 1
            if check == False:
                Facebook.message_error = 'No user found!'
                self.Pre_Messages()
            root = Tk()
            root.attributes('-fullscreen',True)
            Label(root,text='Messages',bg='blue',fg='white',font='Arial 12 bold').place(x=500,y=50)
            Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30,
                                                                                                               y=30)
            if frnd in ID.Name and frnd != self.From:
                post = '1'
                while True:
                    if post == '0':
                        Facebook(ID())
                    check = True
                    i = 1
                    self.N_index = 1
                    self.F_index=1
                    while wb['Messages'].cell(row=i,column=2).value != None:
                        if wb['Messages'].cell(row=i,column=2).value == self.From:
                            if wb['Messages'].cell(row=i,column=3).value == frnd:

                                check = False
                                self.F_index=i
                        if wb['Messages'].cell(row=i,column=3).value == self.From:
                            if wb['Messages'].cell(row=i,column=2).value == frnd:
                                check = False
                                self.F_index=i
                        i+=1
                        self.N_index = i
                    self.colm = 4
                    if check:
                        wb['Messages'].cell(row=self.N_index, column=2, value=self.From)
                        wb['Messages'].cell(row=self.N_index, column=3, value=frnd)
                        wb.save('Facebook.xlsx')
                    else:
                        i=4
                        while wb['Messages'].cell(row=self.F_index, column=i).value != None:
                            if i%2!=0:
                                Label(root,text=wb['Messages'].cell(row=self.F_index, column=i).value,fg='red',font='Arial 10 bold').place(x=280,y=(i*25)+3)
                            else:
                                Label(root,text=wb['Messages'].cell(row=self.F_index, column=i).value,bg='blue',fg='white',font='Arial 10 bold').place(x=255,y=(i*25)+3)
                            i += 1
                            self.colm=i
                    self.enter_message = Entry(root,textvariable=StringVar(),width=30)
                    self.enter_message.place(x=300,y=i*25+20,height=25)
                    Button(root,text='Send',width=4,bg='blue',fg='white',command=self.Command).place(x=490,y=i*25+20)
                    root.mainloop()
    def Command(self):
        frnd = self.search.get()
        frnd_index = ID.Name.index(frnd) + 2
        post = self.enter_message.get()
        if post != '0':
            if self.F_index != 1:
                wb['Messages'].cell(row=self.F_index, column=self.colm, value=self.From)
                wb['Messages'].cell(row=self.F_index, column=self.colm+1, value=post)
                wb['All Users'].cell(row=frnd_index,column=11).value += 1
                wb['All Users'].cell(row=frnd_index, column=12).value += 1
                wb.save('Facebook.xlsx')
            else:
                wb['Messages'].cell(row=self.N_index, column=self.colm, value=self.From)
                wb['Messages'].cell(row=self.N_index, column=self.colm + 1, value=post)
                wb['All Users'].cell(row=frnd_index, column=11).value += 1
                wb['All Users'].cell(row=frnd_index, column=12).value += 1
                wb.save('Facebook.xlsx')
            self.Messaging()
    def Finding_index(self,wb,row1,check,sheet):
        i = 1
        if check == 'C':
            while wb[sheet].cell(row=row1, column=i).value != None:
                i += 1
                col = i
        else:
            while wb[sheet].cell(row=i, column=row1).value != None:
                i += 1
                col = i
        return col
    def MainPage(self):
        def setting():
            Setting.Bio_Changed = ''
            ID.block = ''
            ID.ban = ''
            ID.unban = ''
            ID.unblock = ''
            Setting.setting(Facebook.Sett)
        def post_call():
            ID.Posts(Facebook.id)
        def privacy_call():
            ID.Privacy(Facebook.id)
        Facebook.message_error = ''
        root = Tk()
        Facebook.value = IntVar()
        root.title('facebook')
        root.attributes('-fullscreen',True)
        Label(root,text='facebook',bg='blue',fg='white',font='Arial 12 bold').place(x=20,y=20)
        Button(root,text='Refresh',bg='red',fg='white',command=self.Refresh,font='Arial 9 bold').place(x=600,y=10)
        Button(root,bg='blue',width=10,fg='white',text='Chats',command=self.Pre_Messages,font = 'Simple 12 bold').place(x=260,y=40,height=50)
        Button(root, bg='blue', fg='white',width=10, text='Settings',command=setting,font = 'Simple 12 bold').place(x=390,y=40,height=50)
        Button(root, bg='blue', fg='white', text='Friend Requests',command=self.check_friends,font = 'Simple 12 bold').place(x=520,y=40,height=50)
        Button(root, bg='blue', fg='white',width=10, text='Search',command=self.Pre_Search,font = 'Simple 12 bold').place(x=680,y=40,height=50)
        Button(root, bg='blue', fg='white',text='Notifications',command=self.Notification,font = 'Simple 12 bold').place(x=810,y=40,height=50)
        Button(root, bg='blue', fg='white',text='Post',command=post_call,font = 'Simple 10 bold').place(x=640,y=120,height=30)
        Button(root, bg='blue', fg='white',width=10, text='Log Out', command=self.log_out, font='Simple 12 bold').place(x=940, y=40,height=50)
        Facebook.searched = ''
        Button(root, bg='blue', fg='white', width=10, text='Profile', command=self.Profile,font='Simple 12 bold').place(x=1070, y=40, height=50)
        Button(root, text='Previous', bg='white', fg='blue', width=10, font='Arial 18 bold',command=self.Previous_Post).place(x=200, y=650)
        Button(root, text='Next', bg='blue', fg='white', width=10, font='Arial 18 bold', command=self.Next_Post).place(x=900, y=650)
        Facebook.post = Entry(root,width=30,textvariable=StringVar())
        Facebook.post.insert(0,'Type Something here...')
        def click(event):
            Facebook.post.delete(0,'end')
        Facebook.post.bind('<Button-1>',click)
        Facebook.post.place(x=450,y=120,height=30)
        Checkbutton(root,text='Privacy',onvalue=1,offvalue=0,variable=Facebook.value).place(x=700,y=120)
        Label(root,text=Facebook.posted,fg='green',font='Arial 8 bold').place(x=450,y=160)
        frnd_list = []
        a=3
        while wb['Friends'].cell(row=Facebook.row+1,column=a).value != None:
            frnd_list.append(wb['Friends'].cell(row=Facebook.row+1,column=a).value)
            a+=1
        wb.save('Facebook.xlsx')
        if wb['Posts']['B'+str(self.i)].value in frnd_list or wb['Posts']['B'+str(self.i)].value == Facebook.From:
            owner = wb['Posts']['B' + str(self.i)].value
            owner_ind = ID.Name.index(owner) + 2
            check = True
            if wb['Posts']['A'+str(self.i)].value == 1:
                a = 1
                while wb['Ban_Post'].cell(row=self.i,column=a).value != None:
                    if wb['Ban_Post'].cell(row=self.i,column=a).value == Facebook.From:
                        check = False
                    a+=1
            a = 1
            while wb['Block'].cell(row=owner_ind, column=a).value != None:
                if wb['Block'].cell(row=owner_ind, column=a).value == Facebook.From:
                    check = False
                a += 1
            if wb['Posts']['B' + str(self.i)].value == Facebook.From:
                check= True
            if check:
                Label(root,text=wb['Posts']['B'+str(self.i)].value,fg='blue',font='Amaze 14 bold').place(x=430,y=200)
                Label(root,text=wb['Posts']['C'+str(self.i)].value,fg='red',font='Amaze 13 bold').place(x=510,y=202)
                Label(root,text=wb['Posts']['D'+str(self.i)].value,fg='black',font='Amaze 13 bold').place(x=430,y=240)
                if Facebook.cmnt != 'show':
                    Button(root,text='Comment',bg='blue',fg='white',font='Arial 16 bold',command=self.Show_Comment).place(x=620,y=300)
                else:
                    self.check = 5
                    if Facebook.cmnt == 'show':
                        while wb['Posts'].cell(row=self.i,column=self.check).value != None:
                            if self.check % 2 != 0:
                                Label(root,text=wb['Posts'].cell(row=self.i, column=self.check).value,fg='red',font='Arial 10 bold').place(x=450,y=300 + Facebook.Cmnt_Difference)
                            else:
                                Label(root,text=wb['Posts'].cell(row=self.i, column=self.check).value,fg='blue',font='Arial 10 bold').place(x=500,y=300+ Facebook.Cmnt_Difference)
                                Facebook.Cmnt_Difference += 30
                            self.check += 1
                        self.enter_comment = Entry(root,width=25,textvariable=StringVar())
                        self.enter_comment.place(x=450,y=Facebook.Cmnt_Difference + 300)
                        Button(root,text='Comment',bg='blue',fg='white',font = 'Arial 10 bold',command=self.Save_Coment).place(x=620,y=Facebook.Cmnt_Difference + 300)
                        Facebook.Cmnt_Difference = 0
                        wb.save('Facebook.xlsx')
                root.mainloop()
            else:
                pass
        elif wb['Posts']['B'+str(self.i+1)].value != None:
            self.Next_Post()
        else:
            Label(root,text='Happy Facebook',fg='red',font='Harrington 20 bold').place(x=450,y=300)
            pass
        root.mainloop()
    def Refresh(self):
        self.i = 1
        self.MainPage()
    def log_out(self):
        Facebook.id=id
        Facebook.Sett = set
        root = Tk()
        row = wb['All Users']['J2'].value
        root.title('facebook')
        root.attributes('-fullscreen', True)
        Label(root,text='facebook',bg='blue',fg='white',font='Arial 25 bold').place(x=80,y=170)
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold',command=self.MainPage).place(x=30, y=30)
        # Label(root, text='facebook', bg='blue', fg='white', font='Arial 12 bold').place(x=20, y=20)
        Label(root,text='By Sarib Gardazi',fg='red',font='Arial 12 bold',width=20).place(x=250,y=180)
        Label(root,text='Shaping those who will shape the future',fg='blue',font='Arial 20 bold').place(x=100,y=240)
        frame = Frame(root,bg='#fffcfc')
        frame.place(x=900,y=180,width=400,height=350)
        username = Label(root, text='Username',fg='red',bg='#fffcfc', font='Simple 12 bold')
        password = Label(root, text="Password",fg='red',bg='#fffcfc', font='Simple 12 bold')
        self.name = Entry(root, width=30,bg='#fffcfc', textvariable=StringVar())
        self.name.insert(0,wb['All Users']['B'+str(row+1)].value)
        self.log_Pasword = Entry(root, width=30,bg='#fffcfc', textvariable=StringVar(),show='*')
        username.place(x=950, y=200, height=30)
        password.place(x=950, y=250, height=30)
        self.name.place(x=1050, y=200, height=30)
        self.log_Pasword.place(x=1050, y=250, height=30)
        space = Label(root,text=Facebook.error,fg='red',bg='#fffcfc',font='Arial 12 italic')
        space.place(x=1000,y=280,height=50)
        Button(root, text='Log Out', width=30, bg='blue', fg='#fffcfc', font='Simple 12 bold',
               command=self.Out).place(x=950, y=330, height=50)
        # b1.grid(row=3,column=4)
        root.mainloop()
    def Out(self):
        row = wb['All Users']['J2'].value
        if self.log_Pasword.get() == wb['All Users']['C' + str(row + 1)].value:
            wb['All Users']['J2'] = None
            wb.save('Facebook.xlsx')
            Facebook.error = ''
            Facebook(ID(),Setting())
        else:
            Facebook.error = 'Wrong Password'
            self.log_out()
    def Show_Comment(self):
        Facebook.cmnt = 'show'
        self.MainPage()
    def Save_Coment(self):
        comment = self.enter_comment.get()
        wb['Posts'].cell(row=self.i, column=self.check).value = self.From
        wb['Posts'].cell(row=self.i,column=self.check+1).value = comment
        frnd = wb['Posts']['B'+str(self.i)].value
        frnd_index = ID.Name.index(frnd) + 2
        wb['All Users'].cell(row=frnd_index, column=11).value += 1
        wb['All Users'].cell(row=frnd_index, column=14).value += 1
        wb.save('Facebook.xlsx')
        self.MainPage()
    def Next_Post(self):
        self.i+=1
        Facebook.cmnt = ''
        self.MainPage()
    def Previous_Post(self):
        if self.i > 1:
            self.i-=1
        Facebook.cmnt = ''
        self.MainPage()
    def Next_Page(self):
        Facebook.counter+=1
        Facebook.cmnt = ''
        self.MainPage()
    def Pre_Page(self):
        if Facebook.counter > 1:
            Facebook.counter -= 1
            Facebook.cmnt = ''
        else:
            self.i-=2
            Facebook.cmnt = 1
            Facebook.cmnt = ''
        self.MainPage()
    def Notification(self):
        row=Facebook.row
        col=2
        while wb['Messages'].cell(row=row,column=col).value != None:
            col+=1
        col-=2
        root = Tk()
        root.title('facebook')
        root.attributes('-fullscreen',True)
        # if wb['All Users'].cell(row=Facebook.row+1, column=11).value != 0:
        Label(root, text='Notifications', bg='blue', fg='white', font='Arial 16 bold').place(x=500, y=50)
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=self.MainPage).place(x=30, y=30)
        Label(root,text=wb['All Users'].cell(row=Facebook.row+1, column=13).value,fg='red',font='Arial 14 bold').place(x=350,y=200)
        Label(root,text='New Friend Requests',fg='blue',font='Arial 14 bold').place(x=375,y=200)
        Label(root,text=wb['All Users'].cell(row=Facebook.row+1, column=12).value,fg='red',font='Arial 14 bold').place(x=350,y=230)
        Label(root,text='New Messages',fg='blue',font='Arial 14 bold').place(x=375,y=230)
        Label(root,text=wb['All Users'].cell(row=Facebook.row+1, column=14).value,fg='red',font='Arial 14 bold').place(x=350,y=260)
        Label(root, text='New Comments', fg='blue', font='Arial 14 bold').place(x=375, y=260)
        wb['All Users'].cell(row=Facebook.row+1, column=11).value = 0
        wb['All Users'].cell(row=Facebook.row + 1, column=12).value = 0
        wb['All Users'].cell(row=Facebook.row + 1, column=13).value = 0
        wb['All Users'].cell(row=Facebook.row + 1, column=14).value = 0
        wb.save('Facebook.xlsx')
        root.mainloop()
    def Quit(self):
        self.MainPage()
class ID:
    name2 = ''
    Pasword2 = ''
    df = pd.read_excel('Facebook.xlsx')
    Name = list(df.iloc[:,1])
    scam = '                                  '
    block = ''
    ban = ''
    unban = ''
    unblock = ''
    page_like = ''
    page_dislike =''
    c_page = ''
    Page_Status = ''
    Page_Row = 1
    page_check = ''
    row = 1
    def __init__(self):
        pass
    def Pre_Sign_up(self):
        root = Tk()
        root.title('facebook')
        root.attributes('-fullscreen', True)
        Label(text='facebook', bg='blue', fg='white', font='Arial 25 bold').place(x=1000, y=100)
        frame = Frame(root, bg='#fffcfc')
        frame.place(x=700, y=180, width=400, height=350)
        Label(root,text='facebook',bg='blue',fg='white',font='Arial 25 bold').place(x=80,y=170)
        Label(root,text='By Sarib Gardazi',fg='red',font='Arial 12 bold').place(x=250,y=180)
        Label(root,text='Shaping those who will shape the future',fg='blue',font='Arial 20 bold').place(x=100,y=240)
        username = Label(root, text='Username', fg='black', bg='#fffcfc', font='Simple 12 bold')
        password = Label(root, text="Password", fg='black', bg='#fffcfc', font='Simple 12 bold')
        Resources = Label(root, text='Resources', fg='black', bg='#fffcfc', font='Simple 12 bold')
        self.name = Entry(root, width=30, bg='#fffcfc', textvariable=StringVar())
        self.name.insert(0,'username')
        def click(event):
            self.name.delete(0,'end')
        self.name.bind('<Button-1>',click)
        self.Pasword = Entry(root, width=30, bg='#fffcfc', textvariable=StringVar())
        self.Pasword.insert(0,'Password')
        def click(event):
            self.Pasword.delete(0,'end')
        self.Pasword.bind('<Button-1>',click)
        self.Reso = Entry(root, width=30, bg='#fffcfc', textvariable=StringVar())
        self.Reso.insert(0,'Email or phone number')
        def click(event):
            self.Reso.delete(0,'end')
        self.Reso.bind('<Button-1>',click)
        username.place(x=750, y=200, height=30)
        password.place(x=750, y=250, height=30)
        Resources.place(x=750,y=300, height=30)
        self.name.place(x=850, y=200, height=30)
        self.Pasword.place(x=850, y=250, height=30)
        self.Reso.place(x=850,y=300,height = 30)
        space = Label(root, text=ID.scam, fg='red', bg='#fffcfc', font='Arial 12 italic')
        space.place(x=800, y=350, height=50)
        Label(text='OR', fg='black', bg='white', font='Arial 15 bold').place(x=880, y=390)
        Button(root, text='Create Account', width=27, bg='green', fg='#fffcfc', font='Simple 14 bold',
               command=self.Sign_Up).place(x=750, y=430, height=50)
        root.mainloop()
    def Sign_Up(self):
        self.name2 = self.name.get()
        self.passwor2 = self.Pasword.get()
        check = True
        while check:
            name = self.name.get()
            if name in ID.Name:
                ID.scam = 'This User name is already Selected'
                self.Pre_Sign_up()
            else:
                check = False
        password = self.Pasword.get()
        resource = self.Reso.get()
        self.About = None
        self.Address = None
        self.nickname = None
        self.Friends = []
        n=1
        while wb['All Users']['A'+str(n)].value != None:
            n+=1
        wb['All Users']['A' + str(n)] = n-1
        wb['All Users']['B'+str(n)] = name
        wb['Friend_Requests']['B' + str(n)] = name
        wb['All Users']['C' + str(n)] = password
        wb['All Users']['D' + str(n)] = resource
        wb['All Users']['E' + str(n)] = 'Empty'
        wb['All Users']['F' + str(n)] = 'Empty'
        wb['All Users']['G' + str(n)] = 'Empty'
        wb['All Users']['H' + str(n)] = 'Empty'
        wb['All Users']['I' + str(n)]= 'Empty'
        wb['Friends']['A' + str(n)] = n-1
        wb['Friend_Requests']['A' + str(n)] = n-1
        wb['Messages']['A' + str(n)] = n-1
        wb['Friends']['B'+str(n)] = name
        wb['All Users']['K'+str(n)] = 0
        wb['All Users']['L' + str(n)] = 0
        wb['All Users']['M' + str(n)] = 0
        wb['All Users']['N' + str(n)] = 0
        wb.save('Facebook.xlsx')
        ID.df = pd.read_excel('Facebook.xlsx')
        ID.Name = list(ID.df.iloc[:,1])
        wb.save('Facebook.xlsx')
        index = ID.Name.index(name)
        wb['All Users']['J2'] = index + 1
        wb.save('Facebook.xlsx')
        Facebook.LOG_IN_SIGN(Facebook(ID(),Setting()))
    def Privacy(self):
        return True
    def Posts(self):
        row = ID.Name.index(Facebook.From)+2
        for i in range(1,20000):
            if wb['Posts']['A'+str(i)].value == None:
                max_row = i
                break
        post = Facebook.post.get()
        privacy = Facebook.value.get()
        pList = []
        for i in range(1,2000):
            print(row)
            if wb['Ban'].cell(row=row,column=i).value == None:
                break
            else:
                pList.append(wb['Ban'].cell(row=row,column=i).value)
        if privacy == 1:
            for i in range(len(pList)):
                wb['Ban_Post'].cell(row=max_row,column=i+1).value = pList[i]
        wb['Posts']['A'+str(max_row)] = privacy
        wb['Posts']['B'+str(max_row)] = Facebook.From
        wb['Posts']['C'+str(max_row)] = Time
        wb['Posts']['D' + str(max_row)] = post
        wb.save('Facebook.xlsx')
        if privacy == 1:
            pass
        Facebook.posted = 'Posted Successfully'
        Facebook.MainPage(Facebook(ID(),Setting()))
        wb.save('Facebook.xlsx')
    def Block(self):
        if Setting.block.get() in ID.Name:
            if Setting.block.get() == Facebook.From:
                ID.block = 'You cannot block to yourself'
                Setting.setting(Setting())
            row = wb['All Users']['J2'].value + 1
            test = 1
            while wb['Block'].cell(row=row,column=test).value != None:
                test+=1
            wb['Block'].cell(row=row,column=test).value = Setting.block.get()
            wb.save('Facebook.xlsx')
            ID.block='Blocked Successfully'
            Setting.setting(Setting())
        else:
            ID.block = 'Invalid Username'
            Setting.setting(Setting())
    def Ban(self):
        if Setting.ban.get() in ID.Name:
            if Setting.ban.get() == Facebook.From:
                ID.ban = 'You cannot ban to yourself'
                Setting.setting(Setting())
            row = wb['All Users']['J2'].value + 1
            test = 1
            while wb['Ban'].cell(row=row,column=test).value != None:
                test+=1
            wb['Ban'].cell(row=row,column=test).value = Setting.ban.get()
            wb.save('Facebook.xlsx')
            ID.ban='Banned Successfully'
            Setting.setting(Setting())
        else:
            ID.ban = 'Invalid Username'
            Setting.setting(Setting())
    def Unban(self):
        if Setting.Unban.get() in ID.Name:
            row = wb['All Users']['J2'].value + 1
            test = 1
            check = False
            while wb['Ban'].cell(row=row, column=test).value != None:
                if wb['Ban'].cell(row=row, column=test).value == Setting.Unban.get():
                    check = True
                    Found = test
                test += 1
            if check:
                while wb['Ban'].cell(row=row, column=Found).value != None:
                    wb['Ban'].cell(row=row, column=Found).value = wb['Ban'].cell(row=row, column=Found+1).value
                    wb.save('Facebook.xlsx')
                    Found+=1
                wb.save('Facebook.xlsx')
                ID.unban = 'UnBanned Successfully'
                Setting.setting(Setting())
            ID.unban = 'Not Found'
            Setting.setting(Setting())
        ID.unban = 'Invalid Username'
        Setting.setting(Setting())
        wb.save('Facebook.xlsx')
    def Unblock(self):
        if Setting.Unblock.get() in ID.Name:
            row = wb['All Users']['J2'].value + 1
            test = 1
            check = False
            while wb['Block'].cell(row=row, column=test).value != None:
                if wb['Block'].cell(row=row, column=test).value == Setting.Unblock.get():
                    check = True
                    Found = test
                test += 1
            if check:
                while wb['Block'].cell(row=row, column=Found).value != None:
                    wb['Block'].cell(row=row, column=Found).value = wb['Block'].cell(row=row, column=Found+1).value
                    wb.save('Facebook.xlsx')
                    Found+=1
                wb.save('Facebook.xlsx')
                ID.unblock = 'UnBlocked Successfully'
                Setting.setting(Setting())
            ID.unblock = 'Not found'
            Setting.setting(Setting())
        ID.unblock = 'Invalid Username'
        Setting.setting(Setting())
        wb.save('Facebook.xlsx')
    def Like(self):
        test = 1
        check=False
        while wb['Page'].cell(row=test,column=1).value != None:
            if wb['Page'].cell(row=test,column=1).value == Setting.like.get():
                check = True
                break
            test+=1
        if check:
            colm = 1
            while wb['Page'].cell(row=test,column=colm).value != None:
                if wb['Page'].cell(row=test,column=colm).value == Facebook.From:
                    ID.Page_Status ='Already Liked'
                    self.Search_page()
                    check = False
                    break
                colm +=1
            if check:
                wb['Page'].cell(row=test, column=colm).value = Facebook.From
                wb.save('Facebook.xlsx')
                ID.Page_Status ='Liked Successfully'
                self.Search_page()
        else:
            Page_Status = 'No Page found with this name'
            Setting.Pages(Setting())
    def DisLike(self):
        test = 1
        check = False
        while wb['Page'].cell(row=test, column=1).value != None:
            if wb['Page'].cell(row=test, column=1).value == Setting.like.get():
                check = True
                break
            test += 1
        if check:
            colm = 1
            while wb['Page'].cell(row=test, column=colm).value != None:
                if wb['Page'].cell(row=test, column=colm).value == Facebook.From:
                    check = False
                    break
                colm += 1
            if check == False:
                while wb['Page'].cell(row=test,column = colm).value != None:
                    wb['Page'].cell(row=test, column=colm).value = wb['Page'].cell(row=test,column = colm+1).value
                    colm+=1
                wb.save('Facebook.xlsx')
                ID.Page_Status = 'Disliked Successfully'
                self.Search_page()
            else:
                ID.Page_Status = 'You havent like this page'
                self.Search_page()
        else:
            ID.Page_Status = 'No Page found with this name'
            self.Search_page()
    def Search_page(self):
        test = 1
        check = False
        while wb['Page'].cell(row=test, column=1).value != None:
            if wb['Page'].cell(row=test, column=1).value == Setting.like.get():
                check = True
                break
            test += 1
        if check:
            root =Tk()
            def call():
                Facebook.MainPage(Facebook(ID(), Setting()))
            root.attributes('-fullscreen',True)
            Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=call).place(x=30, y=60)
            Label(root, text='facebook', bg='blue', fg='white', font='Arial 12 bold').place(x=20, y=20)
            Label(root, text=Setting.like.get(), fg='brown', font='Harrington 24 bold').place(x=300, y=100)
            Label(root, text='Creator', fg='black', font='Arial 14 bold').place(x=300, y=150)
            Label(root, text=wb['Page'].cell(row=test, column=2).value, fg='green', font='Harrington 16 bold').place(x=390,y=150)
            Label(root,text=ID.Page_Status,fg='red',font='Arial 10 bold').place(x=310,y=190)
            Button(root,text='Like',bg='blue',fg='white',font = 'Arial 14 bold',width=12,command=self.Like).place(x=460,y=220)
            Button(root, text='DisLike', bg='white', fg='blue', font='Arial 14 bold', width=12,command=self.DisLike).place(x=300, y=220)
            Button(root, bg='blue', fg='white', text='Post', font='Simple 10 bold',command=self.Page_Post).place(x=500,y=280,height=30)
            ID.Ppost = Entry(root, width=30, textvariable=StringVar(),borderwidth=3)
            ID.Ppost.insert(0, 'Type Something here...')
            ID.Ppost.place(x=310, y=280, height=30)
            Button(root,text='Next',bg='blue',fg='white',font='Aril 14 bold',width=8,command=self.Next_Page).place(x=650,y=600)
            Label(root,text=ID.page_check,fg='brown',font='Arial 9 bold').place(x=310,y=260)
            Button(root, text='Previous', bg='white', fg='blue', font='Aril 14 bold', width=8,command=self.Pre_Page).place(x=300, y=600)
            while wb['Page Post'].cell(row=ID.Page_Row,column=1).value != None:
                if wb['Page Post'].cell(row=ID.Page_Row,column=1).value == Setting.like.get():
                    Label(root,text=wb['Page Post'].cell(row=ID.Page_Row,column=1).value,fg='brown',font='Arial 14 bold').place(x=460,y=340)
                    Label(root, text=wb['Page Post'].cell(row=ID.Page_Row, column=2).value, fg='red',font='Arial 14 bold').place(x=620, y=340)
                    Label(root, text=wb['Page Post'].cell(row=ID.Page_Row, column=3).value, fg='blue',font='Arial 14 bold').place(x=460, y=390)
                    break
                ID.Page_Row +=1
            root.mainloop()
        else:
            ID.page_like = 'No Page found with this name'
            ID.Page_Post_Show(ID())
    def Status(self,Name):
        Pages = []
        for i in range(1,2000):
            if wb['Page'].cell(row=i,column=1).value == None:
                break
            else:
                Pages.append(wb['Page'].cell(row=i,column=1).value)
        Page_ind = Pages.index(Name)
        check = False
        for i in range(1,2000):
            if wb['Page'].cell(row=Page_ind+1,column=i).value == None:
                break
            else:
                if wb['Page'].cell(row=Page_ind+1,column=i).value == Facebook.From:
                    check = True
                    break
        return check
    def Page_Post(self):
        Post_ind = 1
        Name = Setting.like.get()
        post = ID.Ppost.get()
        if self.Status(Setting.like.get()):
            for i in range(1,2000):
                if wb['Page Post'].cell(row=i,column=1).value == None:
                    Post_ind = i
                    break
            wb['Page Post'].cell(row=Post_ind,column=1).value = Name
            wb['Page Post'].cell(row=Post_ind,column=2).value = Time
            wb['Page Post'].cell(row=Post_ind,column=3).value = post
            wb.save('Facebook.xlsx')
            ID.page_check = 'Posted Successfully'
            self.Search_page()
        else:
            ID.page_check = 'Please Like our Page First'
            self.Search_page()
    def Next_Page(self):
        ID.Page_Row +=1
        self.Search_page()
    def Pre_Page(self):
        if ID.Page_Row > 1:
            ID.Page_Row -=1
        self.Search_page()
    def Page_Post_Show(self):
        root = Tk()
        root.attributes('-fullscreen',True)
        def call():
            Facebook.MainPage(Facebook(ID(), Setting()))
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=call).place(x=30, y=60)
        Setting.like = Entry(root, textvariable=StringVar(), width=30, borderwidth=3)
        Setting.like.insert(30, 'Search')
        def click(event):
            Setting.like.delete(0,'end')
        Setting.like.bind('<Button-1>',click)
        Setting.like.place(x=300, y=150, height=30)
        Label(root,text=ID.page_like,fg='red',font='Arial 10 bold').place(x=300,y=180)
        frame = Frame(root, bg='grey', borderwidth=5, relief=SUNKEN)
        frame.place(x=830, y=200, width=300, height=370)
        Button(root, text='Search', bg='Blue', fg='white', font='Arial 12 bold',command=lambda: ID.Search_page(ID())).place(x=490, y=150)
        Label(root, text=ID.c_page, fg='red', bg='grey', font='Arial 10 bold').place(x=900, y=450)
        Button(root, text='Create Page', bg='Blue', fg='white', font='Arial 14 bold',command=lambda: ID.Create_Page(ID())).place(x=900, y=500)
        Setting.page = Entry(root, textvariable=StringVar(), width=30, borderwidth=5)
        Setting.page.insert(0, 'Page Name')
        def click(event):
            Setting.page.delete(0,'end')
        Setting.page.bind('<Button-1>',click)
        Setting.page.place(x=870, y=300, height=40)
        Setting.page_pswrd = Entry(root, textvariable=StringVar(), width=30, borderwidth=5)
        Setting.page_pswrd.insert(3, 'Current Password')
        Setting.page_pswrd.place(x=870, y=390, height=40)
        if wb['Page Post'].cell(row=ID.row,column=1).value != None:
            if self.Status(wb['Page Post'].cell(row=ID.row,column=1).value):
                Label(root,text=wb['Page Post'].cell(row=ID.row,column=1).value,fg='brown',font='Harrington 18 bold').place(x=300,y=200)
                Label(root,text=wb['Page Post'].cell(row=ID.row,column=2).value,fg='red',font='Harrington 18 bold').place(x=470,y=200)
                Label(root, text=wb['Page Post'].cell(row=ID.row, column=3).value, fg='blue',font='Harrington 18 bold').place(x=300, y=250)
        else:
            Label(root,text='No Post to show',fg='red',font = 'Harrington 18 bold').place(x=350,y=300)
        Button(root, text='Next', bg='blue', fg='white', font='Aril 14 bold', width=8,command=self.Next_Page_Post).place(x=550, y=600)
        Button(root, text='Previous', bg='white', fg='blue', font='Aril 14 bold', width=8,command=self.Pre_Page_Post).place(x=200, y=600)
        root.mainloop()
    def Next_Page_Post(self):
        ID.row +=1
        self.Page_Post_Show()
    def Pre_Page_Post(self):
        if ID.row > 1:
            ID.row -= 1
            self.Page_Post_Show()
    def Create_Page(self):
        row = wb['All Users']['J2'].value
        if Setting.page_pswrd.get() == wb['All Users']['C' + str(row + 1)].value:
            test = 1
            while wb['Page'].cell(row=test,column=1).value != None:
                test+=1
            wb['Page'].cell(row=test,column=1).value = Setting.page.get()
            wb['Page'].cell(row=test, column=2).value = Facebook.From
            wb.save('Facebook.xlsx')
            ID.c_page = 'Page Created Successfully'
            self.Page_Post_Show()
        else:
            ID.c_page = 'Wrong Password'
            self.Page_Post_Show()
class Setting:
    Bio_Changed = ''
    page = ''
    page_pswrd = ''
    def __init__(self):
        pass
    def setting(self):
        row = wb['All Users']['J2'].value
        root = Tk()
        root.title('facebook')
        root.attributes('-fullscreen', True)
        frame = Frame(root, bg='grey',borderwidth=5,relief=SUNKEN)
        frame.place(x=700, y=150, width=500, height=550)
        def call():
            Facebook.MainPage(Facebook(ID(),Setting()))
        Button(root, text='Quit', bg='red', fg='white', font='Arial 14 bold', command=call).place(x=30, y=60)
        Label(root, text='facebook', bg='blue', fg='white', font='Arial 12 bold').place(x=20, y=20)
        Label(root,text='Edit Info',bg='blue',fg='white',width=15,font='Arial 14 bold').place(x=960,y=100)
        self.username = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.username.insert(0,wb['All Users']['B' + str(row + 1)].value)
        self.username.place(x=950,y=180)
        Label(root,text='User name',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=180)
        self.pswrd = Entry(root,textvariable=StringVar(),width=20,borderwidth=3,show='*')
        self.pswrd.insert(0,wb['All Users']['C' + str(row + 1)].value)
        self.pswrd.place(x=950,y=230)
        Label(root,text='Password',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=230)
        self.about = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.about.insert(0,wb['All Users']['E' + str(row + 1)].value)
        self.about.place(x=950,y=280)
        Label(root,text='About',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=280)
        self.address = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.address.insert(0,wb['All Users']['F' + str(row + 1)].value)
        self.address.place(x=950,y=330)
        Label(root,text='Address',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=330)
        self.nickname = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.nickname.insert(0,wb['All Users']['G' + str(row + 1)].value)
        self.nickname.place(x=950,y=380)
        Label(root,text='Nick Name',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=380)
        self.education = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.education.insert(0,wb['All Users']['H' + str(row + 1)].value)
        self.education.place(x=950,y=430)
        Label(root,text='Education',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=430)
        self.work = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.work.insert(0,wb['All Users']['I' + str(row + 1)].value)
        self.work.place(x=950,y=480)
        Label(root,text='Work',fg='black',bg='grey',font='Arial 14 bold').place(x=840,y=480)
        self.current_password = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        self.current_password.place(x=950,y=550)
        Label(root,text=Setting.Bio_Changed,fg='red',bg='grey',font='Arial 11 italic').place(x=810,y=520)
        Label(root,text='Current Password',fg='blue',bg='grey',font='Arial 14 bold').place(x=750,y=550)
        Button(root,text='Change',bg='blue',fg='white',font='Arial 14 bold',command=self.Change_Bio).place(x=890,y=600)
        Setting.block = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        Setting.block.insert(0,'username')
        Setting.block.place(x=300,y=120)
        Button(root,text='Block',bg='Blue',fg='white',font='Arial 12 bold',command=lambda : ID.Block(ID())).place(x=325,y=150)
        Label(root, text=ID.block, fg='red', font='Arial 10 bold').place(x=300, y=90)
        Setting.ban = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        Setting.ban.insert(0,'username')
        Setting.ban.place(x=300,y=220)
        Button(root,text='Ban',bg='Blue',fg='white',font='Arial 12 bold',command=lambda : ID.Ban(ID())).place(x=335,y=250)
        Label(root, text=ID.ban, fg='red', font='Arial 10 bold').place(x=300, y=190)
        Setting.Unban = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        Setting.Unban.insert(0,'username')
        Setting.Unban.place(x=300,y=320)
        Button(root,text='UnBan',bg='Blue',fg='white',font='Arial 12 bold',command=lambda : ID.Unban(ID())).place(x=325,y=350)
        Label(root,text=ID.unban,fg='red',font='Arial 10 bold').place(x=300,y=290)
        Setting.Unblock = Entry(root,textvariable=StringVar(),width=20,borderwidth=3)
        Setting.Unblock.insert(0,'username')
        Setting.Unblock.place(x=300,y=420)
        Button(root,text='UnBlock',bg='Blue',fg='white',font='Arial 12 bold',command=lambda : ID.Unblock(ID())).place(x=320,y=450)
        Button(root, text='Pages', bg='red', fg='white', font='Arial 16 bold',command= lambda : ID.Page_Post_Show(ID())).place(x=320, y=550)
        Label(root,text=ID.unblock,fg='red',font='Arial 10 bold').place(x=300,y=390)
        root.mainloop()
    def Change_Bio(self):
        row = wb['All Users']['J2'].value
        if self.current_password.get() == wb['All Users']['C' + str(row + 1)].value:
            wb['All Users']['B' + str(row + 1)] = self.username.get()
            wb['All Users']['C' + str(row + 1)] = self.pswrd.get()
            wb['All Users']['E' + str(row + 1)] = self.about.get()
            wb['All Users']['F' + str(row + 1)] = self.address.get()
            wb['All Users']['G' + str(row + 1)] = self.nickname.get()
            wb['All Users']['H' + str(row + 1)] = self.education.get()
            wb['All Users']['I' + str(row + 1)] = self.work.get()
            wb.save('Facebook.xlsx')
            Setting.Bio_Changed = 'Bio Changed Successfully'
            self.setting()
        else:
            Setting.Bio_Changed = 'Invalid Password'
            self.setting()
facebook = Facebook(ID(),Setting())

