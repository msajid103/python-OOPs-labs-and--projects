import openpyxl
from tkinter import *
from tkinter import messagebox
from functools import partial
wb = openpyxl.load_workbook("Facebook.xlsx")
ws = wb["Sheet1"]
ws2 = wb["Sheet2"]
ws3 = wb["Sheet3"]
ws4 = wb["Sheet4"]
ws5 = wb["Sheet5"]
ws6 = wb["Sheet6"]
fix = [ws2,ws3,ws4,ws5,ws6]
for i in fix:
    i['A'][0].value = 'Full_NAME'
header = ['Full_NAME', 'User_name', ' Date_birth', 'Job', "Education", 'Status',
          'Other', 'Password', 'friendNotification',
          'messageNotification'
          '']
for i in range(1, len(header) + 1):
    ws.cell(row=1, column=i).value = header[i - 1]
wb.save('Facebook.xlsx')
class User:
    def __init__(self):
        self.name = None
        self.user_name = None
    def signUp(self,name,username, password,dob,tkWindow):
        self.name = name.get().upper()
        self.user_name = username.get()
        if self.name == 'FULL NAME' or self.user_name =='User Name' or password.get()=='Password' or dob.get()=='Date OF Birth' :
            messagebox.showinfo("", 'Please! Fill All The Entry')
            Functionality.signUp(self, tkWindow)
        elif self.name == '' or self.user_name == '' or password.get() == '' or dob.get() == '':
            messagebox.showinfo("",'Please! Fill All The Entry')
            Functionality.signUp(self,tkWindow)
        else:
            flag = True
            for col in ws['B']:
                if col.value == self.user_name:
                    flag = False
                    messagebox.showinfo("", ' This user alreasdy Exist')
            if flag:
                ws.cell(row=ws.max_row + 1, column=1).value = self.name
                ws2.cell(row=1, column=ws2.max_column + 1).value = self.name
                ws3.cell(row=1, column=ws2.max_column).value = self.name
                ws4.cell(row=1, column=ws2.max_column).value = self.name
                ws5.cell(row=1, column=ws5.max_column+2).value = self.name
                ws.cell(row=ws.max_row, column=2).value = self.user_name
                ws.cell(row=ws.max_row, column=8).value = password.get()
                ws.cell(row=ws.max_row, column=3).value = dob.get()
                ws.cell(row=ws.max_row, column=9).value = False
                ws.cell(row=ws.max_row, column=10).value = False
                ws.cell(row=ws.max_row, column=11).value = False
                wb.save("Facebook.xlsx")
                messagebox.showinfo("","Account SUCCESSFULLY Created")
                tkWindow.destroy()
                Functionality()
    def searchMember(self,name,wind):
        wind.destroy()
        name = name.get().upper()
        print(name)
        window = Tk()
        window.title('Profile')
        window.geometry("400x250")
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self,window))
        window.config(menu=menubar)
        flag = True
        for ro in range(len(ws['A'])):
            if ws['A'][ro].value == name:
                flag = False
                for j in range(1, 8):
                    Label3 = Label(window, text=ws.cell(row=1, column=j).value + ":")
                    Label3.grid(column=0, row=j)
                    Label3 = Label(window, text=ws.cell(row=ro + 1, column=j).value)
                    Label3.grid(column=2, row=j)
        if flag:
            Label(window, text="🙁! Result Not Found").place(x=110,y=90)


            # Functionality.searchmember(self, window)
    def showProfile(self,window):
        window.destroy()
        new_window = Tk()
        new_window.title('Profile')
        new_window.geometry("400x250")
        for j in range(1, 8):
            Label3 = Label(new_window, text=ws.cell(row=1, column=j).value + ":")
            Label3.grid(column=0,row=j)
            Label3 = Label(new_window, text=ws.cell(row=self.i + 1, column=j).value)
            Label3.grid(column=2, row=j)
        menubar = Menu(new_window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self,new_window))
        menubar.add_command(label="Edit Profile", command=lambda: Functionality.editprofile(self,new_window))
        new_window.config(menu=menubar)

class Facebook:
    def __init__(self):
        self.user_name = None
        self.loginStatus = False
        self.name = None
    def login(self,name,paword,tkwindow):
        if name.get()=='name' or name.get()==''or paword.get() == 'password' or paword.get() == '':
            messagebox.showinfo('','Enter! valid entry')
            tkwindow.destroy()
            Functionality()
            return False
        self.user_name = name.get()
        password = paword.get()
        self.loginStatus = False
        for col in range(len(ws['B'])):
            if ws['B'][col].value == self.user_name and ws['H'][col].value == password:
                self.name = ws['A'][col].value
                self.i = col
                Functionality.Button(self,tkwindow)
                self.loginStatus = True
                wb.save("Facebook.xlsx")

        if not self.loginStatus:
            messagebox.showinfo("", "Wrong pass word or user name")
            tkwindow.destroy()
            Functionality()
    def notification(self):
        if not ws['I'][self.i].value and not ws['j'][self.i].value and not ws['K'][self.i].value:
            messagebox.showinfo('Notification','No Notification')
        if ws['I'][self.i].value:
            messagebox.showinfo('Notification', 'You have new friend request')
            ws['I'][self.i].value = False
        if ws['j'][self.i].value:
            messagebox.showinfo('Notification', "You have new message")
            ws['j'][self.i].value = False
        if ws['K'][self.i].value:
            messagebox.showinfo('Notification',"You have new comment on your post")
            ws['K'][self.i].value = False

    def editinfo(self,job, Education,STATUS,Other):
        if job.get()=='' and Education.get()==''and STATUS.get()==''and Other.get()=='':
            messagebox.showinfo('', "Fill Entery")
            return False
        for col in range(len(ws['B'])):
            if ws['B'][col].value == self.user_name:
                ws.cell(row=col + 1, column=4).value = job.get()
                ws.cell(row=col + 1, column=5).value = Education.get()
                ws.cell(row=col + 1, column=6).value = STATUS.get()
                ws.cell(row=col + 1, column=7).value = Other.get()

                wb.save("Facebook.xlsx")
                messagebox.showinfo('',"SUCCESSFUL profile edited")

        # Facebook.functionalty(self)
    def logOut(self,window):
        self.loginStatus = False
        window.destroy()
        Functionality()
    def functionalty(self):
        pre = int(input("\nPress logout = 1 ,Profile = 2"
                        "\nSearch Member = 3,Friend Request = 4"
                        "\n Messages = 5,Post= 6\n"))

        if pre == 1:
            Facebook.logOut(self)
        elif pre == 2:
            inp = input("Press 1 For Show, 2 For Edit Profil : ")
            if inp == '1':
                User.showProfile(self,self.i)
            elif inp == '2':
                Facebook.editinfo(self, self.user_name)
            else:
                pass
        elif pre == 3:
            # Facebook.privateFriend(self)
            User.searchMember(self)
        elif pre == 4:
            inp = input("Press 1 For Send Request, 2 For See and Accept : ")
            if inp == '1':
                Facebook.friendRequest(self)
            elif inp == '2':
                Facebook.acceptrequest(self)
            else:
                pass
        elif pre == 5:
            inp = input("Press 1 For Send , 2 For See Messages : ")
            if inp == '1':
                Facebook.send_message(self)
            elif inp == '2':
                Facebook.vewMessages(self)
            else:
                pass
        elif pre == 6:
            inp = input("Press 1 For Posting, 2 For See And Commenting,\nPost Search by Word = 3: ")
            if inp == '1':
                Facebook.posting(self)
            elif inp == '2':
                Facebook.showpost(self)
            elif inp == '3':
                Facebook.searchByWord(self)
            else:
                pass
        else:
            Facebook.functionalty()
    def vewMessages(self,wind):
        wind.destroy()
        window = Tk()
        window.title('Profile')
        window.geometry("400x250")
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.message(self, window))
        window.config(menu=menubar)
        flag = True
        for i in range(2,ws4.max_row+1):
            if ws4.cell(row= i, column=self.i+1).value != None:
                Label(window,text=ws4.cell(row= i, column=self.i+1).value).grid(column=0, row=i)
                flag =False
        if flag:
            Label(window,text='Empty Message box').place(x=120,y=100)

    def friendRequest(self,name,wind):
        if name.get() == '' or name.get() == 'Enter full name of user':
            messagebox.showinfo('', "Enter! valid name")
            return Functionality.searchmember(self,wind)
        name = name.get().upper()
        flag = True
        if name == self.name:
            messagebox.showinfo('', "You can't Sent yourself")
            flag = False
        for k in range(1,ws2.max_column+2):
            if ws2.cell(row=1, column=k).value == name and k != self.i + 1:
                flag = False
                if Facebook.alreadyfriend(self,k):
                    ws.cell(row=k, column=9).value = True
                    l = 1
                    while l <= ws2.max_row + 3:
                        if ws2.cell(row=l, column=k).value == None:
                            ws2.cell(row=l, column=k).value = self.name
                            messagebox.showinfo('',"Successfully Sent")
                            wb.save("Facebook.xlsx")
                            break
                        l += 1
                k += 1
        if flag:
            messagebox.showinfo('', "This user dos't exist")
        Functionality.searchmember(self,wind)
    def alreadyfriend(self,a):
        for i in range(2,ws.max_row+1):
            if ws2.cell(row=i, column=a).value == self.name:
                messagebox.showinfo('', "Already sent")
                return False
            if ws3.cell(row=i, column=a).value == self.name:
                messagebox.showinfo('', "Already Friend")
                return False
        return True
    def acceptrequest(self,wind):
        l = 1
        while l <= ws2.max_row + 1:
            val = ws2.cell(row=l + 1, column=self.i + 1).value
            if val != None:
                print(f'To accept the Friend request of {val}')
                # pre = input('press 1,2 reject else press any key')
                Facebook.press(self, val, wind)
                if Facebook.press(self, val, wind) == "1":
                    ws2.cell(row=l + 1, column=self.i + 1).value = None
                    for m in range(ws3.max_row + 1):
                        if ws3.cell(row=m + 1, column=self.i + 1).value == None:
                            ws3.cell(row=m + 1, column=self.i + 1).value = val
                            k = 1
                            while k <= ws3.max_column:
                                if ws3.cell(row=1, column=k).value == val:
                                    li = 1
                                    while li <= ws3.max_row + 1:
                                        if ws3.cell(row=li + 1, column=k).value == None:
                                            ws3.cell(row=li + 1, column=k).value = self.name
                                            wb.save("Facebook.xlsx")
                                            messagebox.showinfo('','Accepted')
                                            # print('accepted')
                                            break
                                        li += 1
                                k += 1
                # elif Facebook.press(self, val, wind) == '2':
                #     ws2.cell(row=l + 1, column=self.i + 1).value=None
                #     messagebox.showinfo('', 'Rejected')

                    # print("rejected")
                    wb.save("Facebook.xlsx")

            l += 1
        Facebook.functionalty(self)
    def press(self,val,wind):
        wind.destroy()
        window = Tk()
        window.geometry('400x250')
        Label(window, text=val).place(x=110,y=110)
        Button(window, text="Accept",command=lambda :partial(Functionality.Button(self, window)),bg="Blue", width=36, fg="White"
               , activebackground="Blue").place(x=110, y=170)
        return '1'


    def send_message(self,to,messa,wind):

        To = to.get().upper()
        message = messa.get()
        if To == self.name:
            messagebox.showinfo('',"You can't sent to yourself")
            return Functionality.message(self,wind)
        if message == 'Compose Message'or message == "":
            messagebox.showinfo('',"Please! Enter valid message")
            return Functionality.message(self,wind)
        if to.get() == 'To' or To =="" :
            messagebox.showinfo('',"Please! Enter valid name")
            return Functionality.message(self,wind)

        From = self.name
        flag = True
        for j in range(len(ws4['1'])):
            if ws4['1'][j].value == To:
                flag = False
                k = 1
                while k <= ws4.max_row+1:
                    if ws4.cell(row= k, column=j+1).value == None:
                        ws4.cell(row= k, column=j+1).value = message + "   from  :  " + From
                        ws.cell(row=j+1, column=10).value = True
                        messagebox.showinfo('', 'Message sent')
                        wb.save('Facebook.xlsx')
                        break
                    k += 1
        if flag:
            messagebox.showinfo('', 'This user not exist!')
        Functionality.message(self, wind)
    def posting(self,post,wind):
        user_post_choice = post.get()
        if user_post_choice == ''or user_post_choice == 'Write post here....':
            messagebox.showinfo('','Enter something\n In Post box')
            return Functionality.post(self,wind)
        for i in range(1,len(ws5['1'])):
            if ws5['1'][i].value == self.name:
                k = 1
                while k <= ws5.max_row:
                    if ws5.cell(row=k + 1, column=i+1).value == None:
                        ws5.cell(row=k + 1, column=i+1).value = user_post_choice
                        wb.save("Facebook.xlsx")
                        messagebox.showinfo('',"Successfully Posted ")
                        break
                    k += 1
        Functionality.post(self,wind)
    def showpost(self):
            poster = input("Enter the name of the person see post and comment on it : ").upper()
            flag = True
            for i in range(ws3.max_row+1):
                if ws3.cell(row=i+1, column=self.i+1).value == poster :
                    flag = False
                    for E in range(len(ws5['1'])):
                        if ws5['1'][E].value == poster:
                            Y = 2
                            while Y <= ws5.max_row+1:
                                if ws5.cell(row= Y, column=E+1).value == None:
                                    print("\nNo more post")
                                    Facebook.functionalty(self)
                                print(" ---Post is : ", ws5.cell(row=Y, column=E+1).value)
                                pre = input("TO comment on this post press 1,2 for next post,Enter to exist ")
                                if pre == '1':
                                    Facebook.comment(self,Y,E)
                                elif pre == '':
                                    Facebook.functionalty(self)
                                elif pre == '2':
                                    pass
                                Y += 1
            if flag:
                print(f"You can't see and comment on Post of {poster}")
    def comment(self,Y,E):
        user_comment = input(" Write your comment  : ")
        if user_comment == "":
            Facebook.functionalty(self)
        elif ws5.cell(row=Y, column=E).value != None:
            ws5.cell(row=Y, column=E).value += "[ " + self.name + " commented : " + user_comment + " ] \n"
            ws.cell(row=(E/2)+1, column=11).value = True
        else:
            ws5.cell(row=Y, column=E).value = "[ " + self.name + " commented : " + user_comment + " ]"
            ws.cell(row=E/2
                    , column=11).value = True
        wb.save("Facebook.xlsx")
    def searchByWord(self,word,wind):
        word = word.get()
        window = Tk()
        window.geometry('400x250')
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.post(self, window))
        window.config(menu=menubar)
        if word == "":
            messagebox.showinfo('',"Enter! some word")
            return Functionality.post(self,wind)
        flag = True
        for i in range(1,ws5.max_column+2):
            for r in range(1,ws5.max_row+2):
                if word in str(ws5.cell(row=r, column=i).value):
                    wind.destroy()
                    Label(window,text=ws5.cell(row=1, column=i).value+'Posted :').grid(column=0,row=0)
                    Label(window,text=ws5.cell(row=r,column=i).value).grid(column=1,row=0)
                    flag = False
        if flag:
            wind.destroy()
            Label(window, text="🙁! Result Not Found").place(x=110, y=90)

        window.mainloop()


class Functionality:
    def __init__(self,window=None):
        if window != None:
            window.destroy()
        tkWindow = Tk()
        tkWindow.geometry('400x400')
        tkWindow.title('Login')
        # Label(tkWindow, text="facebook",height=3,width=9,font='sans 30 bold',bg="Blue",fg="White").place(x=60,y=190)
        username = StringVar()
        e = Entry(tkWindow, textvariable=username)
        e.place(x=110,y=20)
        e.insert(0,'username')
        def click(event):
            e.delete(0,'end')
        e.bind('<Button-1>',click)

        password = StringVar()
        e2 = Entry(tkWindow, textvariable=password)
        e2.place(x=110, y=50)
        e2.insert(0, 'password')
        def click(event):
            e2.delete(0, 'end')
        e2.bind('<Button-1>', click)

        validateLogin = partial(Facebook.login,self, username, password,tkWindow)
        Button(tkWindow, text="Log In",font='sans 11 bold', command=validateLogin,
               bg="Blue",fg="White",activebackground="Blue").place(x = 135,y=80)

        Label(tkWindow,text="_________OR_________").place(x = 105,y=120)

        Button(tkWindow, text="Creat New Acount",font='sans 11 bold',
               command=lambda :Functionality.signUp(self,tkWindow),
               bg="Green",fg="White",activebackground="Blue").place(x=95,y=150)
        tkWindow.mainloop()
    def signUp(self,tkwindow):
        tkwindow.destroy()
        Window = Tk()
        Window.geometry('400x300')
        Window.title('Create New acount')
        Label(Window, text="What's your name?",font='sans 11 bold',fg="Black").place(x=100,y=10)
        Label(Window, text="Enter the name you use in real life",fg="Black").place(x=80,y=40)
        name = StringVar()

        e1 = Entry(Window,bg="Gray", textvariable=name)

        e1.place(x=30, y=80)
        e1.insert(0, 'Full Name')
        def click(event):
            e1.delete(0, 'end')
        e1.bind('<Button-1>', click)
        # Entry(Window, textvariable=name).grid(row=0, column=2)
        # Label(Window, text="User Name").grid(row=1, column=0)
        username = StringVar()
        e2 = Entry(Window,bg="Gray", textvariable=username)
        e2.place(x=160, y=80)
        e2.insert(0, 'User Name')

        def click(event):
            e2.delete(0, 'end')

        e2.bind('<Button-1>', click)
        # Entry(Window, textvariable=username).grid(row=1, column=2)
        # Label(Window, text="Password").grid(row=2, column=0)
        password = StringVar()
        e3 = Entry(Window,bg="Gray", textvariable=password)
        e3.place(x=30, y=110)
        e3.insert(0, 'Password')

        def click(event):
            e3.delete(0, 'end')

        e3.bind('<Button-1>', click)
        # Entry(Window, textvariable=password, show='*').grid(row=2, column=2)
        # Label(Window, text="Date OF Birth").grid(row=3, column=0)
        dob = StringVar()
        e4 = Entry(Window,bg="Gray", textvariable=dob)
        e4.place(x=30, y=140)
        e4.insert(0, 'Date OF Birth')

        def click(event):
            e4.delete(0, 'end')

        e4.bind('<Button-1>', click)
        # Entry(Window, textvariable=dob).grid(row=3, column=2)
        signup = partial(User.signUp,self,name,username, password,dob,Window)
        Button(Window, text="Submit", command=signup,bg="Blue",width=36,fg="White"
               ,activebackground="Blue").place(x=40,y=170)
        menubar = Menu(Window)
        menubar.add_command(label="⇐Back", command=lambda:Functionality(Window) )
        Window.config(menu=menubar)
        Window.mainloop()
    def Button(self,window):
        window.destroy()
        Window = Tk()
        Window.geometry('400x300')
        Window.title(self.name)
        name = StringVar()
        e1 = Entry(Window, textvariable=name)
        e1.place(x=40,y=15)
        # e1.place(x=30, y=80)
        e1.insert(0, 'Enter Name To Search ')

        def click(event):
            e1.delete(0, 'end')

        e1.bind('<Button-1>', click)
        Button(Window, text="🔎🔎", command=partial(User.searchMember, self, name, Window),
              fg="Blue",activebackground="Blue").place(x=2,y=10)
        menubar = Menu(Window)
        menubar.add_command(label="Logout", command=lambda :Facebook.logOut(self,Window))
        menubar.add_command(label="Profile", command=lambda : User.showProfile(self,Window))
        menubar.add_command(label="Notification", command=lambda : Facebook.notification(self))
        menubar.add_command(label="SendRequest", command=lambda : Functionality.searchmember(self,Window))
        menubar.add_command(label="Message", command=lambda : Functionality.message(self,Window))
        menubar.add_command(label="Post", command=lambda : Functionality.post(self,Window))
        menubar.add_command(label="Friend", command=lambda : Facebook.acceptrequest(self,Window))
        Window.config(menu=menubar)
        Window.mainloop()

    def editprofile(self,window):
        window.destroy()
        new_window = Tk()
        new_window.title('Edit Profile')
        new_window.geometry("400x250")
        Label(new_window, text="JOB").grid(row=0, column=0)
        job = StringVar()
        Entry(new_window, textvariable=job).grid(row=0, column=2)

        Label(new_window, text="Education").grid(row=1, column=0)
        Education = StringVar()
        Entry(new_window, textvariable=Education).grid(row=1, column=2)

        Label(new_window, text="Status").grid(row=2, column=0)
        STATUS = StringVar()
        Entry(new_window, textvariable=STATUS).grid(row=2, column=2)

        Label(new_window, text="Other").grid(row=3, column=0)
        Other = StringVar()
        Entry(new_window, textvariable=Other).grid(row=3, column=2)
        edit = partial(Facebook.editinfo,self, job, Education,STATUS,Other)
        Button(new_window, text="Submit",font='sans 11 bold', command=edit,
               bg="Blue",fg="White",activebackground="Blue").place(x = 95,y=110)
        menubar = Menu(new_window)
        menubar.add_command(label="⇐Back", command=lambda: User.showProfile(self, new_window))
        new_window.config(menu=menubar)
        new_window.mainloop()
    def searchmember(self,wind):
        wind.destroy()
        window = Tk()
        window.title('SendRequest')
        window.geometry("400x250")

        name = StringVar()
        e1 = Entry(window,  textvariable=name)

        e1.place(x=110, y=60)
        e1.insert(0, 'Enter full name of user')

        def click(event):
            e1.delete(0, 'end')

        e1.bind('<Button-1>', click)
        Button(window, text="Send", font='sans 11 bold', command=partial(Facebook.friendRequest,self,name,window),
               bg="Blue", fg="White", activebackground="Blue").place(x=145, y=100)
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self, window))
        window.config(menu=menubar)
        window.mainloop()
    def message(self,wind):
        wind.destroy()
        window = Tk()
        window.title('Message')
        window.geometry("400x250")


        name = StringVar()
        e1 = Entry(window, textvariable=name)
        e1.place(x=110, y=60)
        e1.insert(0, 'To')

        def click(event):
            e1.delete(0, 'end')

        e1.bind('<Button-1>', click)
        message = StringVar()
        e2 = Entry(window, textvariable=message)

        e2.place(x=110, y=90)
        e2.insert(0, 'Compose Message')

        def click(event):
            e2.delete(0, 'end')

        e2.bind('<Button-1>', click)
        Button(window, text="Send", font='sans 11 bold', command=partial(Facebook.send_message, self, name, message,window),
               bg="Blue", fg="White", activebackground="Blue").place(x=145, y=120)
        Button(window, text="VeiwMessage", font='sans 11 bold',
               command=partial(Facebook.vewMessages,self, window),
               bg="Blue", fg="White", activebackground="Blue").place(x=115, y=160)
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self, window))
        window.config(menu=menubar)
        window.mainloop()

    def post(self, wind):
        wind.destroy()
        window = Tk()
        window.title('Post')
        window.geometry("400x250")
        name = StringVar()
        e = Entry(window, textvariable=name)
        e.place(x=40, y=15)
        # e1.place(x=30, y=80)
        e.insert(0, 'Post Search by word...')
        def click(event):
            e.delete(0, 'end')

        e.bind('<Button-1>', click)
        Button(window, text="🔎🔎", command=partial(Facebook.searchByWord, self, name, window),
               fg="Blue", activebackground="Blue").place(x=2, y=10)

        post = StringVar()
        e1 = Entry(window, textvariable=post)

        e1.place(x=110, y=60)
        e1.insert(0, 'Write post here....')

        def click(event):
            e1.delete(0, 'end')

        e1.bind('<Button-1>', click)
        Button(window, text="Post", font='sans 11 bold', command=partial(Facebook.posting, self, post, window),
               bg="Blue", fg="White", activebackground="Blue").place(x=145, y=100)
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self, window))
        window.config(menu=menubar)
        window.mainloop()



# Facebook()
Functionality()


