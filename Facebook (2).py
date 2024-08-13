import openpyxl
from tkinter import *
from tkinter import messagebox
from functools import partial
wb = openpyxl.load_workbook("Facebook.xlsx")
ws = wb["Sheet1"]
ws2 = wb["Sheet2"]
ws3 = wb["Sheet3"]
ws4 = wb["Sheet4"]
ws5 = wb["Sheet5"]
ws6 = wb["Sheet6"]
fix = [ws2,ws3,ws4,ws5,ws6]
for i in fix:
    i['A'][0].value = 'Full_NAME'
header = ['Full_NAME', 'User_name', ' Date_birth', 'Job', "Education", 'Status',
          'Other', 'Password', 'friendNotification',
          'messageNotification'
          '']
for i in range(1, len(header) + 1):
    ws.cell(row=1, column=i).value = header[i - 1]
wb.save('Facebook.xlsx')
class User:
    def __init__(self):
        self.name = None
        self.user_name = None
    def signUp(self,name,username, password,dob,tkWindow):
        self.name = name.get().upper()
        self.user_name = username.get()
        if self.name == '' or self.user_name =='' or password.get()==''or dob.get()=='' :
            messagebox.showinfo("",'Please! Fill All The Entry')
            Functionality.signUp(self,tkWindow)

        else:
            flag = True
            for col in ws['B']:
                if col.value == self.user_name:
                    flag = False
                    messagebox.showinfo("", ' This user alreasdy Exist')
            if flag:
                ws.cell(row=ws.max_row + 1, column=1).value = self.name
                ws2.cell(row=1, column=ws2.max_column + 1).value = self.name
                ws3.cell(row=1, column=ws2.max_column).value = self.name
                ws4.cell(row=1, column=ws2.max_column).value = self.name
                ws5.cell(row=1, column=ws5.max_column+2).value = self.name
                ws.cell(row=ws.max_row, column=2).value = self.user_name
                ws.cell(row=ws.max_row, column=8).value = password.get()
                ws.cell(row=ws.max_row, column=3).value = dob.get()
                ws.cell(row=ws.max_row, column=9).value = False
                ws.cell(row=ws.max_row, column=10).value = False
                ws.cell(row=ws.max_row, column=11).value = False
                wb.save("Facebook.xlsx")
                messagebox.showinfo("","Account SUCCESSFULLY Created")
                tkWindow.destroy()
                Functionality()
    def searchMember(self,name,wind):
        wind.destroy()
        name = name.get().upper()
        print(name)
        window = Tk()
        window.title('Profile')
        window.geometry("400x250")
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.searchmember(self,window))
        window.config(menu=menubar)
        flag = True
        for ro in range(len(ws['A'])):
            if ws['A'][ro].value == name:
                flag = False
                for j in range(1, 8):
                    Label3 = Label(window, text=ws.cell(row=1, column=j).value + ":")
                    Label3.grid(column=0, row=j)
                    Label3 = Label(window, text=ws.cell(row=ro + 1, column=j).value)
                    Label3.grid(column=2, row=j)
        if flag:
            Label3 = Label(window, text="🙁! Result Not Found").place(x=110,y=90)


            # Functionality.searchmember(self, window)
    def showProfile(self,window):
        window.destroy()
        new_window = Tk()
        new_window.title('Profile')
        new_window.geometry("400x250")
        for j in range(1, 8):
            Label3 = Label(new_window, text=ws.cell(row=1, column=j).value + ":")
            Label3.grid(column=0,row=j)
            Label3 = Label(new_window, text=ws.cell(row=self.i + 1, column=j).value)
            Label3.grid(column=2, row=j)
        menubar = Menu(new_window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self,new_window))
        menubar.add_command(label="Edit Profile", command=lambda: Functionality.editprofile(self,new_window))
        new_window.config(menu=menubar)

class Facebook:
    def __init__(self):
        self.user_name = None
        self.loginStatus = False
        self.name = None
    def login(self,name,paword,tkwindow):
        self.user_name = name.get()
        password = paword.get()
        self.loginStatus = False
        for col in range(len(ws['B'])):
            if ws['B'][col].value == self.user_name and ws['H'][col].value == password:
                self.name = ws['A'][col].value
                self.i = col
                Functionality.Button(self,tkwindow)
                self.loginStatus = True
                Facebook.notification(self)
                wb.save("Facebook.xlsx")

        if not self.loginStatus:
            messagebox.showinfo("", "Wrong pass word or user name")
            tkwindow.destroy()
            Functionality()
    def notification(self):
        if ws['I'][self.i].value:
            print("You have new friend request")
            ws['I'][self.i].value = False
        if ws['j'][self.i].value:
            print("You have new message")
            ws['j'][self.i].value = False
        if ws['K'][self.i].value:
            print("You have new comment on your post")
            ws['K'][self.i].value = False
    def editinfo(self,job, Education,STATUS,Other):
        for col in range(len(ws['B'])):
            if ws['B'][col].value == self.user_name:
                ws.cell(row=col + 1, column=4).value = job.get()
                ws.cell(row=col + 1, column=5).value = Education.get()
                ws.cell(row=col + 1, column=6).value = STATUS.get()
                ws.cell(row=col + 1, column=7).value = Other.get()

                wb.save("Facebook.xlsx")
                messagebox.showinfo('',"SUCCESSFUL profile edited")

        # Facebook.functionalty(self)
    def logOut(self,window):
        self.loginStatus = False
        window.destroy()
        Functionality()
    def functionalty(self):
        pre = int(input("\nPress logout = 1 ,Profile = 2"
                        "\nSearch Member = 3,Friend Request = 4"
                        "\n Messages = 5,Post= 6\n"))

        if pre == 1:
            Facebook.logOut(self)
        elif pre == 2:
            inp = input("Press 1 For Show, 2 For Edit Profil : ")
            if inp == '1':
                User.showProfile(self,self.i)
            elif inp == '2':
                Facebook.editinfo(self, self.user_name)
            else:
                pass
        elif pre == 3:
            # Facebook.privateFriend(self)
            User.searchMember(self)
        elif pre == 4:
            inp = input("Press 1 For Send Request, 2 For See and Accept : ")
            if inp == '1':
                Facebook.friendRequest(self)
            elif inp == '2':
                Facebook.acceptrequest(self)
            else:
                pass
        elif pre == 5:
            inp = input("Press 1 For Send , 2 For See Messages : ")
            if inp == '1':
                Facebook.send_message(self)
            elif inp == '2':
                Facebook.vewMessages(self)
            else:
                pass
        elif pre == 6:
            inp = input("Press 1 For Posting, 2 For See And Commenting,\nPost Search by Word = 3: ")
            if inp == '1':
                Facebook.posting(self)
            elif inp == '2':
                Facebook.showpost(self)
            elif inp == '3':
                Facebook.searchByWord(self)
            else:
                pass
        else:
            Facebook.functionalty()
    def vewMessages(self):
        for i in range(2,ws4.max_row+1):
            print(ws4.cell(row= i, column=self.i+1).value)
        Facebook.functionalty(self)
    def friendRequest(self):
        name = input("Write the name of user:").upper()
        flag = True
        for k in range(1,ws2.max_column+2):
            if ws2.cell(row=1, column=k).value == name and k != self.i + 1:
                if Facebook.alreadyfriend(self,k):
                    ws.cell(row=k, column=9).value = True
                    flag = False
                    l = 1
                    while l <= ws2.max_row + 3:
                        if ws2.cell(row=l, column=k).value == None:
                            ws2.cell(row=l, column=k).value = self.name
                            print("Sended")
                            wb.save("Facebook.xlsx")
                            break
                        l += 1
                else:
                    print("you are already send friend request")
                k += 1
        if flag:
            print("This user dos't exist")
        Facebook.functionalty(self)
    def alreadyfriend(self,a):
        for i in range(2,ws.max_row+1):
            if ws2.cell(row=i, column=a).value == self.name or ws3.cell(row=i, column=a).value == self.name:
                return False
        return True
    def acceptrequest(self):
        l = 1
        while l <= ws2.max_row + 1:
            val = ws2.cell(row=l + 1, column=self.i + 1).value
            if val != None:
                print(f'To accept the Friend request of {val}')
                pre = input('press 1,2 reject else press any key')
                if pre == '1':
                    ws2.cell(row=l + 1, column=self.i + 1).value = None
                    for m in range(ws3.max_row + 1):
                        if ws3.cell(row=m + 1, column=self.i + 1).value == None:
                            ws3.cell(row=m + 1, column=self.i + 1).value = val
                            k = 1
                            while k <= ws3.max_column:
                                if ws3.cell(row=1, column=k).value == val:
                                    li = 1
                                    while li <= ws3.max_row + 1:
                                        if ws3.cell(row=li + 1, column=k).value == None:
                                            ws3.cell(row=li + 1, column=k).value = self.name
                                            wb.save("Facebook.xlsx")
                                            print('accepted')
                                            break
                                        li += 1
                                k += 1
                elif pre == '2':
                    ws2.cell(row=l + 1, column=self.i + 1).value=None
                    print("rejected")
                    wb.save("Facebook.xlsx")

            l += 1
        Facebook.functionalty(self)
    def send_message(self):
        flag = True
        message = str(input(" Type message ........ "))
        if message == '':
            print("Please Enter valid message")
            return Facebook.functionalty(self)
        To = input(" TO : ").upper()
        From = self.name
        for j in range(len(ws4['1'])):
            if ws4['1'][j].value == To:
                flag = False
                k = 1
                while k <= ws4.max_row+1:
                    if ws4.cell(row= k, column=j+1).value == None:
                        ws4.cell(row= k, column=j+1).value = message + "   from  :  " + From
                        ws.cell(row=j+1, column=10).value = True
                        wb.save('Facebook.xlsx')
                        print('Message sended')
                        break
                    k += 1

        if flag:
            print(" ************ This user not exist! ********** ")
        Facebook.functionalty(self)
    def posting(self):
        print(" Welcome ", self.name, " !  To Posting-------  ")
        for i in range(1,len(ws5['1'])):
            if ws5['1'][i].value == self.name:
                k = 1
                while k <= ws5.max_row:
                    if ws5.cell(row=k + 1, column=i+1).value == None:
                        user_post_choice = input(" Write a post : ")
                        if user_post_choice == '':
                            print('Enter something\nNot Posted')
                            break
                        ws5.cell(row=k + 1, column=i+1).value = user_post_choice
                        wb.save("Facebook.xlsx")
                        print("Successfully Posted ")
                        break
                    k += 1
        Facebook.functionalty(self)
    def showpost(self):
            poster = input("Enter the name of the person see post and comment on it : ").upper()
            flag = True
            for i in range(ws3.max_row+1):
                if ws3.cell(row=i+1, column=self.i+1).value == poster :
                    flag = False
                    for E in range(len(ws5['1'])):
                        if ws5['1'][E].value == poster:
                            Y = 2
                            while Y <= ws5.max_row+1:
                                if ws5.cell(row= Y, column=E+1).value == None:
                                    print("\nNo more post")
                                    Facebook.functionalty(self)
                                print(" ---Post is : ", ws5.cell(row=Y, column=E+1).value)
                                pre = input("TO comment on this post press 1,2 for next post,Enter to exist ")
                                if pre == '1':
                                    Facebook.comment(self,Y,E)
                                elif pre == '':
                                    Facebook.functionalty(self)
                                elif pre == '2':
                                    pass
                                Y += 1
            if flag:
                print(f"You can't see and comment on Post of {poster}")
    def comment(self,Y,E):
        user_comment = input(" Write your comment  : ")
        if user_comment == "":
            Facebook.functionalty(self)
        elif ws5.cell(row=Y, column=E).value != None:
            ws5.cell(row=Y, column=E).value += "[ " + self.name + " commented : " + user_comment + " ] \n"
            ws.cell(row=(E/2)+1, column=11).value = True
        else:
            ws5.cell(row=Y, column=E).value = "[ " + self.name + " commented : " + user_comment + " ]"
            ws.cell(row=E/2
                    , column=11).value = True
        wb.save("Facebook.xlsx")
    def searchByWord(self):
        word = input('Write Word')
        for i in range(1,ws5.max_column+2):

            for r in range(1,ws5.max_row+2):
                if word in str(ws5.cell(row=r, column=i).value):
                    print(f"{ws5.cell(row=1, column=i).value} posted:{ws5.cell(row=r,column=i).value}")

class Functionality:
    def __init__(self,window=None):
        if window != None:
            window.destroy()
        tkWindow = Tk()
        tkWindow.geometry('400x300')
        tkWindow.title('Login')
        Label(tkWindow, text="User Name").grid(row=0, column=0)
        username = StringVar()
        Entry(tkWindow, textvariable=username).grid(row=0, column=2)
        Label(tkWindow, text="Password").grid(row=1, column=0)
        password = StringVar()
        Entry(tkWindow, textvariable=password, show='*').grid(row=1, column=2)
        validateLogin = partial(Facebook.login,self, username, password,tkWindow)
        Button(tkWindow, text="Log In",font='sans 11 bold', command=validateLogin,
               bg="Blue",fg="White",activebackground="Blue").place(x = 95,y=50)

        Label(tkWindow,text="_________OR_________").place(x = 65,y=90)

        Button(tkWindow, text="Creat New Acount",font='sans 11 bold',
               command=lambda :Functionality.signUp(self,tkWindow),
               bg="Green",fg="White",activebackground="Blue").place(x=55,y=120)
        tkWindow.mainloop()
    def signUp(self,tkwindow):
        tkwindow.destroy()
        Window = Tk()
        Window.geometry('400x300')
        Window.title('Create New acount')
        Label(Window, text="Full Name").grid(row=0, column=0)
        name = StringVar()
        Entry(Window, textvariable=name).grid(row=0, column=2)
        Label(Window, text="User Name").grid(row=1, column=0)
        username = StringVar()
        Entry(Window, textvariable=username).grid(row=1, column=2)
        Label(Window, text="Password").grid(row=2, column=0)
        password = StringVar()
        Entry(Window, textvariable=password, show='*').grid(row=2, column=2)
        Label(Window, text="Date OF Birth").grid(row=3, column=0)
        dob = StringVar()
        Entry(Window, textvariable=dob).grid(row=3, column=2)
        signup = partial(User.signUp,self,name,username, password,dob,Window)
        Button(Window, text="Submit", command=signup,bg="Blue",fg="White"
               ,activebackground="Blue").grid(row=4, column=2)
        menubar = Menu(Window)
        menubar.add_command(label="⇐Back", command=lambda:Functionality(Window) )
        Window.config(menu=menubar)
        Window.mainloop()
    def Button(self,window):
        window.destroy()
        Window = Tk()
        Window.geometry('400x300')

        Window.title(self.name)
        Button(Window, text="Logout", command=lambda :Facebook.logOut(self,Window),
                bg="Blue",fg="White",activebackground="Blue").grid(row=4, column=2)
        Button(Window, text="Profile", command=lambda : User.showProfile(self,Window)
, bg="Blue",fg="White",activebackground="Blue").grid(row=4, column=3)

        menubar = Menu(Window)
        menubar.add_command(label="Logout", command=lambda :Facebook.logOut(self,Window))
        menubar.add_command(label="Profile", command=lambda : User.showProfile(self,Window))
        menubar.add_command(label="Search", command=lambda : Functionality.searchmember(self,Window))
        Window.config(menu=menubar)
        Window.mainloop()

    def editprofile(self,window):
        window.destroy()
        new_window = Tk()
        new_window.title('Edit Profile')
        new_window.geometry("400x250")
        Label(new_window, text="JOB").grid(row=0, column=0)
        job = StringVar()
        Entry(new_window, textvariable=job).grid(row=0, column=2)

        Label(new_window, text="Education").grid(row=1, column=0)
        Education = StringVar()
        Entry(new_window, textvariable=Education).grid(row=1, column=2)

        Label(new_window, text="Status").grid(row=2, column=0)
        STATUS = StringVar()
        Entry(new_window, textvariable=STATUS).grid(row=2, column=2)

        Label(new_window, text="Other").grid(row=3, column=0)
        Other = StringVar()
        Entry(new_window, textvariable=Other).grid(row=3, column=2)

        edit = partial(Facebook.editinfo,self, job, Education,STATUS,Other)
        Button(new_window, text="Submit",font='sans 11 bold', command=edit,
               bg="Blue",fg="White",activebackground="Blue").place(x = 95,y=110)

        menubar = Menu(new_window)
        menubar.add_command(label="⇐Back", command=lambda: User.showProfile(self, new_window))
        new_window.config(menu=menubar)
        new_window.mainloop()
    def searchmember(self,wind):
        wind.destroy()
        window = Tk()
        window.title('Search')
        window.geometry("400x250")
        Label(window, text="Search").grid(row=0, column=0)
        name = StringVar()
        Entry(window, textvariable=name).grid(row=0, column=2)
        Button(window, text="Search", font='sans 11 bold', command=partial(User.searchMember,self,name,window),
               bg="Blue", fg="White", activebackground="Blue").place(x=95, y=110)
        menubar = Menu(window)
        menubar.add_command(label="⇐Back", command=lambda: Functionality.Button(self, window))
        window.config(menu=menubar)
        window.mainloop()



# Facebook()
Functionality()

